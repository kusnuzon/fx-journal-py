import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
import tkinter.messagebox as msgbox
from idlelib.tooltip import Hovertip
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import time
from PIL import Image, ImageTk
from PIL.ImageGrab import grabclipboard
import tempfile
import sys

columns = ['Time', 'Symbol', 'Order', 'Risk', 'Category', 'MindState', 'Entry Analysis', 'Management']

############# User Defined Options #######################################################
sheetfilename = 'trading_journal'
categories = ['Med Prob', 'High Prob', 'Low Prob']  # Customize the levels
mindstate = ['Normal', 'Good', 'Bad']              # Customize the levels
row_height_pts = 200      # Height of cell rows
width_info_column = 40    # Width of columns with extended text
#############################################################################################

class Panel(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Trading Journal Manager MOD v1.4")
        self.geometry("+50+50")
        self.resizable(0, 0)
        style = ttk.Style(self)
        style.configure('TNotebook.Tab', padding=(7, 6, 7, 0))
        estyle = ttk.Style()
        estyle.configure("EntryStyle.TEntry", background='white')

        # Variables
        self.sheetfilename = sheetfilename + '.xlsx'
        self.logging_dict = {key: tk.StringVar() for key in columns}
        self.spreadsheet = Spreadsheet(self.sheetfilename)
        self.opentrades, self.opentrade_rownum = self.get_open_trades()
        self.open_trade_id = tk.StringVar()
        self.num_open_trades = tk.StringVar()
        self.num_open_trades.set(f'You have {sum([el != "None" for el in self.opentrades])} open trades')
        self.use_screenshot = tk.IntVar(value=0)  # Default to off since we're using clipboard
        self.new_screenshot = tk.IntVar(value=0)  # Default to off for open trades
        self.clipboard_image = None  # Store clipboard image for preview
        self.temp_image_path = None  # Store path to temporary image file

        # UI setup
        self.label = ttk.Label(self, textvariable=self.num_open_trades, font=('Serif Bold', 20), foreground='black', anchor=tk.CENTER)
        self.label.grid(sticky='news', columnspan=5, pady=5)
        n = ttk.Notebook(self)
        tab_1 = ttk.Frame(n)
        tab_2 = ttk.Frame(n)
        tab_3 = ttk.Frame(n)
        n.add(tab_1, text=" New Trade")
        n.add(tab_2, text=" Open Trades")
        n.add(tab_3, text=" Options")
        n.grid(sticky=tk.NSEW, columnspan=5)

        # Tab 1: New Trades
        ttk.Label(tab_1, text="Symbol:").grid(row=0, column=0, sticky='news', padx=(5, 0), pady=10)
        ttk.Entry(tab_1, textvariable=self.logging_dict['Symbol'], style='EntryStyle.TEntry', width=7).grid(row=0, column=1, sticky='news', padx=(1, 5), pady=10)
        order = ttk.OptionMenu(tab_1, self.logging_dict['Order'], "Buy", "Buy", "Sell")
        order.grid(row=0, column=2, padx=3, pady=10, sticky='news')
        order.config(width=4)
        ttk.Label(tab_1, text="Risk:").grid(row=0, column=3, sticky='news', padx=(3, 0), pady=10)
        risk_entry = ttk.Entry(tab_1, textvariable=self.logging_dict['Risk'], style='EntryStyle.TEntry', width=2)
        risk_entry.grid(row=0, column=4, sticky='news', padx=0, pady=10)
        risk_entry.insert(0, "1")
        ttk.Label(tab_1, text="%").grid(row=0, column=5, sticky='news', padx=0, pady=10)
        ttk.Label(tab_1, text="Category:").grid(row=1, column=0, sticky='news', padx=(5, 0), pady=0)
        ttk.Label(tab_1, text="Mind State:").grid(row=1, column=2, sticky='news', padx=(5, 0), pady=0)
        ttk.OptionMenu(tab_1, self.logging_dict['Category'], categories[0], *categories).grid(row=2, column=0, columnspan=2, padx=5, pady=0, sticky='news')
        ttk.OptionMenu(tab_1, self.logging_dict['MindState'], mindstate[0], *mindstate).grid(row=2, column=2, columnspan=2, padx=5, pady=0, sticky='news')
        ttk.Label(tab_1, text="Entry Analysis:").grid(row=3, column=0, columnspan=2, sticky='news', padx=5, pady=1)
        self.analysis = tk.Text(tab_1, highlightcolor="LightSteelBlue2", width=width_info_column, height=10)
        self.analysis.grid(row=4, column=0, columnspan=5, sticky='news', padx=5, pady=1)
        ttk.Label(tab_1, text="Management Rules:").grid(row=5, column=0, columnspan=2, sticky='news', padx=5, pady=1)
        self.mgmt = tk.Text(tab_1, highlightcolor="LightSteelBlue2", width=width_info_column, height=10)
        self.mgmt.grid(row=6, column=0, columnspan=5, sticky='news', padx=5, pady=1)
        style.configure('my.TButton', foreground='blue')
        ttk.Button(tab_1, style='my.TButton', text='Add Entry', command=self.add_entry).grid(padx=5, pady=1, sticky='news', column=1, columnspan=2)
        ttk.Button(tab_1, style='my.TButton', text='Paste Screenshot', command=lambda: self.paste_screenshot(is_new_trade=True)).grid(padx=5, pady=1, sticky='news', column=0, columnspan=2)

        # Tab 2: Open Trades
        ttk.Label(tab_2, text="Select:").grid(row=0, column=0, sticky='news', padx=(5, 0), pady=10)
        self.openMenu = ttk.OptionMenu(tab_2, self.open_trade_id, self.opentrades[0], *self.opentrades)
        self.openMenu.grid(row=0, column=1, padx=3, pady=10, sticky='news', columnspan=3)
        self.open_trade_id.trace("w", self.OptionMenu_SelectionEvent)
        ttk.Label(tab_2, text="Result/Comments:").grid(row=1, column=0, columnspan=2, sticky='news', padx=5, pady=5)
        self.result = tk.Text(tab_2, highlightcolor="LightSteelBlue2", width=width_info_column, height=10)
        self.result.grid(row=2, column=0, columnspan=5, sticky='news', padx=5, pady=1)
        ttk.Button(tab_2, style='my.TButton', text='Update Open Trade', command=self.add_resulttext).grid(padx=5, pady=1, sticky='news', row=3, column=0, columnspan=2)
        ttk.Button(tab_2, style='my.TButton', text='Paste Screenshot', command=lambda: self.paste_screenshot(is_new_trade=False)).grid(padx=5, pady=1, sticky='news', row=4, column=0, columnspan=2)
        s2 = ttk.Style()
        s2.configure('my2.TButton', foreground='blue', font=('Sans', 12, 'bold'))
        ttk.Button(tab_2, style='my2.TButton', text='Close Trade', command=self.close_opentrade).grid(padx=5, pady=20, sticky='news', row=5, column=2, columnspan=2)

        # Tab 3: Options (simplified, no folder selection needed)
        ttk.Label(tab_3, text="Instructions:", font=('Sans', 10, 'bold')).grid(row=0, column=0, sticky='news', padx=(30, 5), pady=(20, 1))
        ttk.Label(tab_3, text="1. Take a screenshot (e.g., PrtSc or Snipping Tool).\n2. Click 'Paste Screenshot' in New Trade or Open Trades tab.\n3. Preview and confirm to attach the screenshot.\n4. New Trade screenshots save to 'Chart Before'; Open Trades screenshots save to 'Chart After'.").grid(row=1, column=0, sticky='news', padx=(30, 5), pady=5, columnspan=2)

    def paste_screenshot(self, is_new_trade):
        """Capture screenshot from clipboard and show preview."""
        try:
            # Capture image from clipboard
            self.clipboard_image = grabclipboard()
            if not self.clipboard_image:
                tk.messagebox.showerror("Error", "No image found in clipboard.\nPlease take a screenshot first.")
                return

            # Save to temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
                self.temp_image_path = temp_file.name
                self.clipboard_image.save(self.temp_image_path, 'PNG')

            # Show preview
            self.show_preview(is_new_trade)

        except Exception as e:
            tk.messagebox.showerror("Error", f"Failed to paste screenshot: {str(e)}")

    def show_preview(self, is_new_trade):
        """Display a preview of the pasted screenshot."""
        preview_window = tk.Toplevel(self)
        preview_window.title("Screenshot Preview")
        preview_window.geometry("600x400")
        preview_window.resizable(0, 0)

        # Resize image for preview
        img = self.clipboard_image.copy()
        img.thumbnail((500, 300))  # Resize to fit window
        photo = ImageTk.PhotoImage(img)

        # Display image
        label = ttk.Label(preview_window, image=photo)
        label.image = photo  # Keep a reference
        label.pack(pady=10)

        # Buttons
        ttk.Button(preview_window, text="Confirm", command=lambda: self.confirm_screenshot(is_new_trade, preview_window)).pack(side=tk.LEFT, padx=10, pady=10)
        ttk.Button(preview_window, text="Cancel", command=lambda: self.cancel_screenshot(preview_window)).pack(side=tk.RIGHT, padx=10, pady=10)

    def confirm_screenshot(self, is_new_trade, preview_window):
        """Confirm the screenshot and proceed with adding it."""
        preview_window.destroy()
        if is_new_trade:
            self.use_screenshot.set(1)  # Enable screenshot for new trade
        else:
            self.new_screenshot.set(1)  # Enable screenshot for open trade
        tk.messagebox.showinfo("Success", "Screenshot confirmed and ready to attach.")

    def cancel_screenshot(self, preview_window):
        """Cancel the screenshot and clean up."""
        preview_window.destroy()
        if self.temp_image_path and os.path.exists(self.temp_image_path):
            os.unlink(self.temp_image_path)  # Delete temporary file
        self.clipboard_image = None
        self.temp_image_path = None
        self.use_screenshot.set(0)
        self.new_screenshot.set(0)
        tk.messagebox.showinfo("Cancelled", "Screenshot attachment cancelled.")

    def add_entry(self):
        """Add a new trade entry with optional screenshot in Chart Before."""
        if self.use_screenshot.get() and not self.temp_image_path:
            tk.messagebox.showerror("Error", "No screenshot pasted.\nClick 'Paste Screenshot' to add an image.")
            return

        now = time.strftime('%a, %b-%d %H:%M')
        self.logging_dict['Entry Analysis'].set(self.analysis.get("1.0", 'end-1c'))
        self.logging_dict['Management'].set(self.mgmt.get("1.0", 'end-1c'))
        self.logging_dict['Time'].set(now)
        if any([self.logging_dict[key].get() == "" for key in self.logging_dict]):
            tk.messagebox.showerror("Error", "Please fill all fields.")
            return
        self.spreadsheet.add_entry(self.logging_dict, self.use_screenshot.get(), self.temp_image_path)
        self.analysis.delete("1.0", "end")
        self.mgmt.delete("1.0", "end")
        self.cleanup_screenshot()
        self.update_panel()

    def add_resulttext(self):
        """Update an open trade with result text and optional screenshot in Chart After."""
        if self.open_trade_id.get() == 'None':
            return
        if self.new_screenshot.get() and not self.temp_image_path:
            tk.messagebox.showerror("Error", "No screenshot pasted.\nClick 'Paste Screenshot' to add an image.")
            return
        text = self.result.get("1.0", 'end-1c')
        self.spreadsheet.add_text(text, self.opencellrow, 11, align_horiz='left')
        if self.new_screenshot.get():
            self.spreadsheet.add_chart(self.opencellrow, self.temp_image_path, chart_type='after')
        self.cleanup_screenshot()
        self.update_panel()

    def cleanup_screenshot(self):
        """Clean up temporary screenshot data."""
        if self.temp_image_path and os.path.exists(self.temp_image_path):
            os.unlink(self.temp_image_path)
        self.clipboard_image = None
        self.temp_image_path = None
        self.use_screenshot.set(0)
        self.new_screenshot.set(0)

    def get_open_trades(self):
        sh = self.spreadsheet.sheet
        maxrow = self.spreadsheet.numrows
        opentrades = []
        row_number = []
        if maxrow < 2:
            return ['None'], [0]
        for row in range(2, maxrow + 1):
            time, symbol, order, closed = (
                sh.cell(row=row, column=1).value,
                sh.cell(row=row, column=2).value,
                sh.cell(row=row, column=3).value,
                sh.cell(row=row, column=12).value
            )
            this_trade = f"{time} {symbol} {order}"
            if closed == 'X':
                continue
            row_number.append(row)
            opentrades.append(this_trade)
        return (opentrades, row_number) if opentrades else (['None'], [0])

    def close_opentrade(self):
        if self.open_trade_id.get() == 'None':
            return
        self.spreadsheet.add_text('X', self.opencellrow, 12, align_horiz='center')
        self.update_panel()

    def OptionMenu_SelectionEvent(self, *args):
        index_ = self.opentrades.index(self.open_trade_id.get())
        self.opencellrow = self.opentrade_rownum[index_]
        self.result.delete("1.0", "end")
        self.result.insert("1.0", self.spreadsheet.get_text(self.opencellrow, 11))

    def update_panel(self):
        self.opentrades, self.opentrade_rownum = self.get_open_trades()
        self.openMenu.set_menu(self.opentrades[0], *self.opentrades)
        self.num_open_trades.set(f'You have {sum([el != "None" for el in self.opentrades])} open trades')
        self.update()

class Spreadsheet:
    def __init__(self, sheetfilename):
        self.filepath = os.getcwd()
        self.fullpath = os.path.join(self.filepath, sheetfilename)
        try:
            self.workbook = self.get_workbook()
            self.update()
        except PermissionError:
            tk.messagebox.showerror(
                "File Access Error",
                f"The file '{self.fullpath}' is currently open in another application (e.g., Excel).\nPlease close the file and restart the application."
            )
            sys.exit(1)

    def show_file_access_error(self):
        """Display error message for file access issues."""
        tk.messagebox.showerror(
            "File Access Error",
            f"The file '{self.fullpath}' is currently open in another application (e.g., Excel).\nPlease close the file and try again."
        )

    def get_workbook(self):
        """Load or create the workbook, handling file access errors."""
        file_exists = os.path.exists(self.fullpath)
        if file_exists:
            try:
                return openpyxl.load_workbook(self.fullpath)
            except PermissionError:
                self.show_file_access_error()
                raise
        else:
            newbook = openpyxl.Workbook()
            newbook = self.make_sheet_header(newbook)
            return newbook

    def make_sheet_header(self, book):
        """Create the header for a new spreadsheet."""
        sheet = book.active
        columns_ = columns + ['Chart Before', 'Chart After', 'Result/Comments', 'Closed']
        for ix, col in enumerate(columns_):
            sheet.cell(row=1, column=ix + 1).value = col
            sheet.cell(row=1, column=ix + 1).font = Font(bold=True)
            sheet.cell(row=1, column=ix + 1).alignment = Alignment(vertical="center", horizontal="center")
            sheet.column_dimensions[get_column_letter(ix + 1)].width = len(col) * 1.4
        for col in ['Entry Analysis', 'Management', 'Result/Comments']:
            indx = columns_.index(col) + 1
            sheet.column_dimensions[get_column_letter(indx)].width = width_info_column
        sheet.column_dimensions[get_column_letter(columns_.index('Time') + 1)].width = 16
        sheet.column_dimensions[get_column_letter(columns_.index('Chart Before') + 1)].width = width_info_column + 10
        sheet.column_dimensions[get_column_letter(columns_.index('Chart After') + 1)].width = width_info_column + 10
        for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(start_color='FF27E85B', end_color='FF27E85B', fill_type="solid")
        rd = sheet.row_dimensions[1]
        rd.height = 20
        try:
            book.save(self.fullpath)
            return openpyxl.load_workbook(self.fullpath)
        except PermissionError:
            self.show_file_access_error()
            raise

    def add_entry(self, logging_dict, use_screenshot, img_path):
        """Add a new entry to the spreadsheet with optional screenshot in Chart Before."""
        for ix, key in enumerate(logging_dict):
            self.sheet.cell(row=self.numrows + 1, column=ix + 1).value = logging_dict[key].get()
            self.sheet.cell(row=self.numrows + 1, column=ix + 1).alignment = Alignment(vertical="top", wrapText=True)
        rd = self.sheet.row_dimensions[self.numrows + 1]
        rd.height = row_height_pts
        if use_screenshot and img_path:
            self.add_chart(self.numrows + 1, img_path, chart_type='before')
        else:
            self.update()

    def add_chart(self, row, img_path, chart_type='before'):
        """Attach the screenshot to the specified row in Chart Before or Chart After."""
        try:
            if not os.path.exists(img_path):
                tk.messagebox.showerror("Error", f"Image file not found: {img_path}")
                return
            img = openpyxl.drawing.image.Image(img_path)
            # Set anchor based on chart_type
            column_index = 9 if chart_type == 'before' else 10  # Chart Before: col 9, Chart After: col 10
            img.anchor = self.sheet.cell(row=row, column=column_index).coordinate
            img.width = img.width * (row_height_pts / img.height) * 1.333 * 0.95
            img.height = row_height_pts * 1.333 * 0.95
            self.sheet.add_image(img)
            self.update()
        except Exception as e:
            tk.messagebox.showerror("Error", f"Failed to attach screenshot: {str(e)}")

    def add_text(self, text, row, col, align_horiz="left"):
        self.sheet.cell(row=row, column=col).value = text
        self.sheet.cell(row=row, column=col).alignment = Alignment(vertical="top", horizontal=align_horiz, wrapText=True)
        self.update()

    def get_text(self, row, col):
        if row < 1:
            return ""
        text = self.sheet.cell(row=row, column=col).value
        return text if text is not None else ""

    def update(self):
        """Save and reload the workbook, handling file access errors."""
        try:
            self.workbook.save(self.fullpath)
            self.workbook = openpyxl.load_workbook(self.fullpath)
            self.sheet = self.workbook.active
            self.numrows = self.sheet.max_row
        except PermissionError:
            self.show_file_access_error()
            raise

if __name__ == '__main__':
    Panel = Panel()
    Panel.mainloop()
